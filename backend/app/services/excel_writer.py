"""
Excel导出引擎 - 生成14个Sheet的美化Excel报告

Sheet顺序：
1. 数据目录      2. 缺失清单      3. ASIN总览面板
4. 流量结构树    5. 12层动作表    6. 今日执行清单
7. 7天监控计划   8. 否词清单      9. 加词清单
10. 拆精准清单   11. 广告位建议   12. SBV视频诊断
13. LLM诊断报告  14. 原始数据汇总
"""
import os
from pathlib import Path
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.services.utils import ensure_dir


# 颜色定义
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)

# 高亮色
GREEN_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
RED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
BLUE_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FEFCE8", end_color="FEFCE8", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin', color='E5E7EB'),
    right=Side(style='thin', color='E5E7EB'),
    top=Side(style='thin', color='E5E7EB'),
    bottom=Side(style='thin', color='E5E7EB'),
)

WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical='center')
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')


def _style_header(ws, num_cols: int):
    """应用表头样式"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 8, max_width: int = 40):
    """自动调整列宽"""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = min_width
        for row_idx in range(1, min(ws.max_row + 1, 100)):
            cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
            if len(cell_value) > max_len:
                max_len = len(cell_value)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def _add_filters(ws):
    """添加自动筛选"""
    if ws.max_row > 1 and ws.max_column > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def _freeze_header(ws):
    """冻结首行"""
    ws.freeze_panes = 'A2'


def _write_data(ws, headers: List[str], rows: List[Dict[str, Any]], highlight_col: str = None):
    """写入表头和数据，并应用基础格式"""
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGNMENT

    _style_header(ws, len(headers))
    _auto_width(ws)
    _add_filters(ws)
    _freeze_header(ws)


class ExcelReportWriter:
    """Excel报告写入器"""

    def __init__(self, output_dir: str = "outputs"):
        self.output_dir = ensure_dir(output_dir)

    def write_full_report(
        self,
        data_directory: List[Dict],
        missing_reports: List[Dict],
        asin_overview: List[Dict],
        traffic_tree: List[Dict],
        action_table: List[Dict],
        today_tasks: List[Dict],
        monitor_plan: List[Dict],
        negative_keywords: List[Dict],
        add_keywords: List[Dict],
        exact_split: List[Dict],
        placement_suggestions: List[Dict],
        video_diagnosis: List[Dict],
        llm_report: str = "",
        raw_data: List[Dict] = None,
        filename_prefix: str = "广告分析报告",
    ) -> str:
        """
        生成完整14 Sheet报告
        返回: 输出文件路径
        """
        wb = Workbook()
        wb.remove(wb.active)

        # === Sheet 1: 数据目录 ===
        ws1 = wb.create_sheet("数据目录")
        headers1 = ["批次", "运营", "店铺", "站点", "文件名", "Sheet名", "报表类型",
                     "识别置信度", "时间范围", "ASIN数量", "广告类型", "活动数量",
                     "广告组数量", "关键词数量", "ASIN Target数量", "读取状态", "备注"]
        _write_data(ws1, headers1, data_directory)

        # === Sheet 2: 缺失清单 ===
        ws2 = wb.create_sheet("缺失清单")
        headers2 = ["运营", "店铺", "ASIN", "缺失数据", "影响模块", "影响结论",
                     "是否阻断", "降级处理方式", "需要补充的文件"]
        _write_data(ws2, headers2, missing_reports)

        # === Sheet 3: ASIN总览面板 ===
        ws3 = wb.create_sheet("ASIN总览面板")
        headers3 = ["ASIN", "父ASIN", "标题", "Spend", "Sales", "Orders", "ACOS",
                     "TACOS", "Impressions", "Clicks", "CTR", "CVR", "CPC",
                     "Sessions", "Page Views", "Unit Session %", "Buy Box %",
                     "退款率", "B2B订单", "B2B销售额", "页面优先判断", "核心问题", "可信度"]
        _write_data(ws3, headers3, asin_overview)

        # === Sheet 4: 流量结构树 ===
        ws4 = wb.create_sheet("流量结构树")
        headers4 = ["ASIN", "词层级", "目标词/搜索词", "来源", "意图层级",
                     "Impr", "Click", "Spend", "Orders", "Sales", "CPC", "CTR",
                     "CVR", "ACOS", "ROAS", "ABA排名", "月度搜索量", "处理建议"]
        _write_data(ws4, headers4, traffic_tree)

        # === Sheet 5: 12层动作表 ===
        ws5 = wb.create_sheet("12层动作表")
        headers5 = ["优先级", "运营", "店铺", "ASIN", "广告类型", "广告活动", "广告组",
                     "目标词/ASIN", "动作层级", "当前数据", "建议动作", "调整前",
                     "调整后", "调整值", "原因", "预期影响", "执行时间", "负责人确认", "风险标记"]
        _write_data(ws5, headers5, action_table)

        # 动作表特殊高亮
        for row_idx in range(2, len(action_table) + 2):
            action = ws5.cell(row=row_idx, column=11).value or ''
            if '预算' in str(action):
                if '增加' in str(action):
                    for c in range(1, len(headers5) + 1):
                        ws5.cell(row=row_idx, column=c).fill = GREEN_FILL
                elif '减少' in str(action):
                    for c in range(1, len(headers5) + 1):
                        ws5.cell(row=row_idx, column=c).fill = RED_FILL
            elif '否定' in str(action):
                for c in range(1, len(headers5) + 1):
                    ws5.cell(row=row_idx, column=c).fill = ORANGE_FILL
            elif '精准' in str(action) or '拆出' in str(action):
                for c in range(1, len(headers5) + 1):
                    ws5.cell(row=row_idx, column=c).fill = BLUE_FILL

        # === Sheet 6: 今日执行清单 ===
        ws6 = wb.create_sheet("今日执行清单")
        headers6 = ["序号", "优先级", "运营", "店铺", "ASIN", "广告活动", "广告组",
                     "目标词/ASIN", "具体动作", "调整值", "执行人", "执行状态",
                     "截止时间", "备注"]
        _write_data(ws6, headers6, today_tasks)

        # === Sheet 7: 7天监控计划 ===
        ws7 = wb.create_sheet("7天监控计划")
        headers7 = ["目标", "监控指标", "阈值/Trigger", "触发动作", "时间窗口",
                     "负责角色", "复盘日期"]
        _write_data(ws7, headers7, monitor_plan)

        # === Sheet 8: 否词清单 ===
        ws8 = wb.create_sheet("否词清单")
        headers8 = ["ASIN", "广告活动", "广告组", "搜索词", "点击", "订单", "花费",
                     "销售额", "ACOS", "相关性判断", "否定类型", "是否需要负责人确认", "原因"]
        _write_data(ws8, headers8, negative_keywords)

        for row_idx in range(2, len(negative_keywords) + 2):
            for c in range(1, len(headers8) + 1):
                ws8.cell(row=row_idx, column=c).fill = ORANGE_FILL

        # === Sheet 9: 加词清单 ===
        ws9 = wb.create_sheet("加词清单")
        headers9 = ["ASIN", "来源", "新增词", "匹配方式", "建议活动", "建议广告组",
                     "建议初始出价", "意图层级", "原因"]
        _write_data(ws9, headers9, add_keywords)

        # === Sheet 10: 拆精准清单 ===
        ws10 = wb.create_sheet("拆精准清单")
        headers10 = ["ASIN", "原活动", "原广告组", "搜索词", "当前表现",
                      "新活动名称", "新广告组名称", "建议预算", "建议出价",
                      "原组否定动作", "原因"]
        _write_data(ws10, headers10, exact_split)

        for row_idx in range(2, len(exact_split) + 2):
            for c in range(1, len(headers10) + 1):
                ws10.cell(row=row_idx, column=c).fill = BLUE_FILL

        # === Sheet 11: 广告位建议 ===
        ws11 = wb.create_sheet("广告位建议")
        headers11 = ["ASIN", "广告活动", "广告组", "Placement", "Spend", "Sales",
                      "Orders", "ACOS", "CVR", "Top位展示份额", "建议溢价", "原因"]
        _write_data(ws11, headers11, placement_suggestions)

        # === Sheet 12: SBV视频诊断 ===
        ws12 = wb.create_sheet("SBV视频诊断")
        headers12 = ["ASIN", "广告活动", "5秒观看次数", "25%观看", "50%观看",
                      "75%观看", "100%观看", "5秒观看率", "VTR", "vCTR",
                      "问题判断", "优化建议", "前5秒脚本"]
        _write_data(ws12, headers12, video_diagnosis)

        # === Sheet 13: LLM诊断报告 ===
        ws13 = wb.create_sheet("LLM诊断报告")
        llm_headers = ["模块", "内容"]
        _write_data(ws13, llm_headers, [{"模块": "LLM诊断报告", "内容": llm_report}])

        # === Sheet 14: 原始数据汇总 ===
        if raw_data:
            ws14 = wb.create_sheet("原始数据汇总")
            if raw_data:
                raw_headers = list(raw_data[0].keys())
                _write_data(ws14, raw_headers, raw_data)

        # 保存
        import datetime
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{filename_prefix}_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath

    def write_simple_report(self, sheets: Dict[str, tuple]) -> str:
        """
        简单模式：直接传入 {Sheet名: (表头列表, 数据列表)}
        返回: 文件路径
        """
        wb = Workbook()
        wb.remove(wb.active)

        for sheet_name, (headers, rows) in sheets.items():
            ws = wb.create_sheet(sheet_name[:31])
            _write_data(ws, headers, rows)

        import datetime
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"报告_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        return filepath
